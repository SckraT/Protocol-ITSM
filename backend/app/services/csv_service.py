"""
Сервис экспорта и импорта данных: CSV и XLSX.
Полностью портирован из app.py v1 с сохранением формата файлов.
"""
import csv
import io
from datetime import date as date_type

from fastapi import HTTPException, UploadFile
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.executor import Executor
from app.models.item import Item
from app.models.status import Status
from app.repositories.executor_repository import ExecutorRepository
from app.repositories.item_repository import ItemRepository
from app.schemas.common import PriorityEnum, StateEnum
from app.utils.constants import (
    CSV_FIELDS,
    PRIORITY_FILL_COLOR,
    PRIORITY_LABEL_RU,
    STATE_FILL_COLOR,
    STATE_LABEL_RU,
)

# Обратные карты «русская метка → значение enum» — чтобы импорт принимал и значения
# (in_progress), и подписи из экспорта (В работе). Неизвестное → дефолт/None.
_STATE_BY_LABEL = {label: value for value, label in STATE_LABEL_RU.items()}
_PRIORITY_BY_LABEL = {label: value for value, label in PRIORITY_LABEL_RU.items()}
_STATE_VALUES = {e.value for e in StateEnum}
_PRIORITY_VALUES = {e.value for e in PriorityEnum}


def _parse_state(raw: str | None) -> str:
    """Нормализовать состояние из CSV. Неизвестное → in_progress."""
    v = (raw or "").strip()
    if v in _STATE_VALUES:
        return v
    return _STATE_BY_LABEL.get(v, StateEnum.in_progress.value)


def _parse_priority(raw: str | None) -> str | None:
    """Нормализовать приоритет из CSV. Неизвестное → None."""
    v = (raw or "").strip()
    if v in _PRIORITY_VALUES:
        return v
    return _PRIORITY_BY_LABEL.get(v)


def _parse_date(raw: str | None) -> date_type | None:
    """Разобрать дату ISO (YYYY-MM-DD). Невалидное/пустое → None."""
    v = (raw or "").strip()
    if not v:
        return None
    try:
        return date_type.fromisoformat(v)
    except ValueError:
        return None


class CsvService:
    """Сервис экспорта и импорта данных."""

    def __init__(self, session: AsyncSession) -> None:
        self.session = session
        self.item_repo = ItemRepository(session)
        self.exec_repo = ExecutorRepository(session)

    def _executor_str(self, executors: list) -> str:
        """Форматирует список исполнителей: 'Отдел — Имя | Имя2'."""
        parts = []
        for e in executors:
            dept = getattr(e, "department", None)
            dept_name = dept.name if dept else None
            if dept_name:
                parts.append(f"{dept_name} — {e.name}")
            else:
                parts.append(e.name)
        return " | ".join(parts)

    async def export_csv(self) -> bytes:
        """
        Экспортирует все задачи в CSV с UTF-8 BOM.
        Формат исполнителей: 'Отдел — Имя | Имя2' — совместим с v1.
        """
        # Загружаем все задачи с предзагрузкой
        result = await self.session.execute(
            select(Item)
            .options(
                selectinload(Item.executors).selectinload(Executor.department),
                selectinload(Item.statuses),
            )
            .order_by(Item.id)
        )
        items = result.scalars().all()

        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=CSV_FIELDS)
        writer.writeheader()

        for item in items:
            # Последний статус — первый после сортировки DESC
            last_status = None
            if item.statuses:
                last_status = sorted(
                    item.statuses,
                    key=lambda s: (s.status_date or date_type.min, s.id),
                    reverse=True,
                )[0]

            writer.writerow({
                "id": item.id,
                "topic": item.topic,
                "ticket": item.ticket or "",
                "priority": item.priority or "",
                "state": item.state,
                "due_date": str(item.due_date) if item.due_date else "",
                "executors": self._executor_str(item.executors),
                "last_status_date": str(last_status.status_date) if last_status and last_status.status_date else "",
                "last_status_note": last_status.status_note or "" if last_status else "",
            })

        # UTF-8 BOM для корректного открытия в Excel
        content = ("﻿" + output.getvalue()).encode("utf-8")
        return content

    async def export_xlsx(self) -> bytes:
        """
        Экспортирует задачи в Excel (.xlsx) с форматированием.
        Стили полностью повторяют v1.
        """
        # Загружаем все задачи
        result = await self.session.execute(
            select(Item)
            .options(
                selectinload(Item.executors).selectinload(Executor.department),
                selectinload(Item.statuses),
            )
            .order_by(Item.id)
        )
        items = result.scalars().all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Протокол"

        # ── Стили ──────────────────────────────────────────────────────────────
        thin = Side(style="thin", color="D1D5DB")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="1E3A5F")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # ── Заголовки ──────────────────────────────────────────────────────────
        headers = ["№", "Тема", "Тикет", "Приоритет", "Состояние", "Срок",
                   "Исполнители", "Дата статуса", "Последний статус"]
        col_widths = [5, 45, 12, 12, 12, 12, 30, 14, 50]

        for col_idx, (hdr, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col_idx, value=hdr)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"

        center = Alignment(horizontal="center", vertical="top")
        wrap = Alignment(vertical="top", wrap_text=True)
        top = Alignment(vertical="top")

        for row_idx, item in enumerate(items, 2):
            # Последний статус
            last_status = None
            if item.statuses:
                last_status = sorted(
                    item.statuses,
                    key=lambda s: (s.status_date or date_type.min, s.id),
                    reverse=True,
                )[0]

            prio = item.priority or ""
            state = item.state
            exec_str = self._executor_str(item.executors)

            row_data = [
                row_idx - 1,
                item.topic,
                item.ticket or "",
                PRIORITY_LABEL_RU.get(prio, prio),
                STATE_LABEL_RU.get(state, state),
                str(item.due_date) if item.due_date else "",
                exec_str,
                str(last_status.status_date) if last_status and last_status.status_date else "",
                last_status.status_note or "" if last_status else "",
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                if col_idx in (2, 7, 9):
                    cell.alignment = wrap
                elif col_idx in (1, 3, 4, 5, 6, 8):
                    cell.alignment = center
                else:
                    cell.alignment = top

            # Цвет по приоритету и состоянию
            if prio in PRIORITY_FILL_COLOR:
                ws.cell(row=row_idx, column=4).fill = PatternFill("solid", fgColor=PRIORITY_FILL_COLOR[prio])
            if state in STATE_FILL_COLOR:
                ws.cell(row=row_idx, column=5).fill = PatternFill("solid", fgColor=STATE_FILL_COLOR[state])

            ws.row_dimensions[row_idx].height = 20

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    async def import_csv(self, file: UploadFile) -> dict:
        """
        Импортирует задачи из CSV.
        Исполнители сопоставляются по имени; при отсутствии — создаются новые.
        Формат строки исполнителей: 'Отдел — Имя | Имя2' (совместим с v1).
        """
        raw = await file.read()
        try:
            text = raw.decode("utf-8-sig")  # utf-8-sig убирает BOM
        except UnicodeDecodeError:
            raise HTTPException(status_code=400, detail="Файл должен быть в кодировке UTF-8")

        reader = csv.DictReader(io.StringIO(text))
        imported = 0

        for row in reader:
            topic = (row.get("topic") or "").strip()
            if not topic:
                continue

            # Разбираем исполнителей из строки вида "Отдел — Имя | Имя2"
            exec_ids: list[int] = []
            exec_str = (row.get("executors") or "").strip()
            if exec_str:
                for part in exec_str.split("|"):
                    part = part.strip()
                    if not part:
                        continue
                    # Берём часть после " — " (имя) или всю строку
                    name = part.split(" — ")[-1].strip()
                    if not name:
                        continue

                    # Ищем исполнителя по имени, при отсутствии — создаём
                    existing = await self.exec_repo.get_by_name(name)
                    if existing:
                        exec_ids.append(existing.id)
                    else:
                        new_exec = Executor(name=name)
                        created_exec = await self.exec_repo.create(new_exec)
                        exec_ids.append(created_exec.id)

            # Создаём задачу (значения нормализуем — иначе мусор сломает чтение)
            item = Item(
                topic=topic,
                ticket=row.get("ticket") or None,
                priority=_parse_priority(row.get("priority")),
                state=_parse_state(row.get("state")),
                due_date=_parse_date(row.get("due_date")),
            )
            created_item = await self.item_repo.create(item)

            # Привязываем исполнителей
            if exec_ids:
                await self.item_repo.update_executors(created_item, exec_ids)

            # Создаём статус если есть
            s_note = (row.get("last_status_note") or "").strip()
            if s_note:
                status = Status(
                    item_id=created_item.id,
                    status_date=_parse_date(row.get("last_status_date")),
                    status_note=s_note,
                )
                self.session.add(status)
                await self.session.flush()

            imported += 1

        return {"imported": imported}
