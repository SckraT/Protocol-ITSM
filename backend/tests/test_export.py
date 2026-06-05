"""Тесты экспорта: защита от CSV/XLSX formula injection."""
import io

import pytest
from openpyxl import load_workbook


@pytest.mark.asyncio
async def test_csv_export_sanitizes_formula(client):
    """Опасное значение topic экранируется префиксом ' в CSV."""
    await client.post("/api/items", json={"topic": "=cmd|'/c calc'!A0", "state": "in_progress"})
    r = await client.get("/api/export/csv")
    assert r.status_code == 200
    text = r.content.decode("utf-8-sig")
    # Ячейка не должна начинаться с '=' — должна быть экранирована апострофом
    assert "'=cmd" in text
    assert "\n=cmd" not in text and not text.startswith("=cmd")


@pytest.mark.asyncio
async def test_csv_export_keeps_normal_values(client):
    """Обычные значения не изменяются (без лишнего апострофа)."""
    await client.post("/api/items", json={"topic": "Обычная задача", "state": "in_progress"})
    r = await client.get("/api/export/csv")
    text = r.content.decode("utf-8-sig")
    assert "Обычная задача" in text
    assert "'Обычная задача" not in text


@pytest.mark.asyncio
async def test_xlsx_export_sanitizes_formula(client):
    """Опасное значение topic экранируется в XLSX."""
    await client.post("/api/items", json={"topic": "=HYPERLINK(\"http://evil\")", "state": "in_progress"})
    r = await client.get("/api/export/xlsx")
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb.active
    # Колонка «Тема» = 2; строка данных = 2
    topic_cell = ws.cell(row=2, column=2).value
    assert isinstance(topic_cell, str)
    assert topic_cell.startswith("'=HYPERLINK")
