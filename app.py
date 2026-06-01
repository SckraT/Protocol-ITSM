"""
Протокол совещания — бэкенд на FastAPI.

Таблицы:
  - items       — задачи протокола;
  - statuses    — история статусов задачи (N:1 с items);
  - departments — справочник отделов;
  - executors   — справочник исполнителей (с привязкой к отделу).

Исполнители задачи хранятся как JSON-массив ID в items.executors_json.
"""

from fastapi import FastAPI, HTTPException, File, UploadFile
from fastapi.staticfiles import StaticFiles
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional
import sqlite3
import json
import csv
import io
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DB_PATH = os.environ.get("DB_PATH", "data/protocol.db")

app = FastAPI(title="Protocol")


def get_db():
    """Открывает соединение с SQLite с доступом к полям по имени."""
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    """Создаёт таблицы и выполняет накопительные миграции при старте."""
    conn = get_db()

    conn.execute("""
        CREATE TABLE IF NOT EXISTS departments (
            id   INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE
        )
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS executors (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            name          TEXT NOT NULL UNIQUE,
            department_id INTEGER REFERENCES departments(id) ON DELETE SET NULL
        )
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS items (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            topic          TEXT NOT NULL,
            ticket         TEXT,
            priority       TEXT,
            state          TEXT NOT NULL DEFAULT 'in_progress',
            due_date       TEXT,
            executors_json TEXT,
            created_at     TEXT DEFAULT (date('now'))
        )
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS statuses (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            item_id     INTEGER NOT NULL REFERENCES items(id) ON DELETE CASCADE,
            status_date TEXT,
            status_note TEXT,
            created_at  TEXT DEFAULT (datetime('now'))
        )
    """)

    items_cols = [r[1] for r in conn.execute("PRAGMA table_info(items)").fetchall()]
    exec_cols  = [r[1] for r in conn.execute("PRAGMA table_info(executors)").fetchall()]

    # Миграция: старая схема хранила статус прямо в items.
    if "status_date" in items_cols or "status_note" in items_cols:
        rows = conn.execute("SELECT id, status_date, status_note FROM items").fetchall()
        for row in rows:
            if row["status_date"] or row["status_note"]:
                conn.execute(
                    "INSERT INTO statuses (item_id, status_date, status_note) VALUES (?, ?, ?)",
                    (row["id"], row["status_date"], row["status_note"]),
                )
        conn.execute("ALTER TABLE items RENAME TO items_old")
        conn.execute("""
            CREATE TABLE items (
                id             INTEGER PRIMARY KEY AUTOINCREMENT,
                topic          TEXT NOT NULL,
                ticket         TEXT,
                priority       TEXT,
                state          TEXT NOT NULL DEFAULT 'in_progress',
                due_date       TEXT,
                executors_json TEXT,
                created_at     TEXT DEFAULT (date('now'))
            )
        """)
        conn.execute("""
            INSERT INTO items (id, topic, state, due_date, created_at)
            SELECT id, topic, state, due_date, created_at FROM items_old
        """)
        conn.execute("DROP TABLE items_old")
        items_cols = [r[1] for r in conn.execute("PRAGMA table_info(items)").fetchall()]

    # Миграция: добавляем отсутствующие колонки в items.
    for col, dfn in [("ticket", "TEXT"), ("priority", "TEXT"), ("executors_json", "TEXT")]:
        if col not in items_cols:
            conn.execute(f"ALTER TABLE items ADD COLUMN {col} {dfn}")

    # Миграция: переносим старые executor/executor2 в executors_json (имена → имена).
    items_cols = [r[1] for r in conn.execute("PRAGMA table_info(items)").fetchall()]
    if "executor" in items_cols:
        rows = conn.execute("SELECT id, executor, executor2, executors_json FROM items").fetchall()
        for row in rows:
            if row["executors_json"]:
                continue
            names = [n for n in [row["executor"], row["executor2"]] if n]
            if names:
                conn.execute("UPDATE items SET executors_json = ? WHERE id = ?",
                             (json.dumps(names, ensure_ascii=False), row["id"]))

    # Миграция: добавляем department_id к executors (если его нет).
    if "department_id" not in exec_cols:
        conn.execute("ALTER TABLE executors ADD COLUMN department_id INTEGER REFERENCES departments(id) ON DELETE SET NULL")

    # Миграция: executors_json в items — переводим строки-имена в целые ID.
    # Если первый элемент массива — строка, значит данные ещё не мигрированы.
    rows = conn.execute("SELECT id, executors_json FROM items WHERE executors_json IS NOT NULL").fetchall()
    for row in rows:
        try:
            data = json.loads(row["executors_json"])
            if not data or not isinstance(data[0], str):
                continue  # пусто или уже ID
            ids = []
            for name in data:
                name = name.strip()
                if not name:
                    continue
                existing = conn.execute("SELECT id FROM executors WHERE name = ?", (name,)).fetchone()
                if existing:
                    ids.append(existing["id"])
                else:
                    cur = conn.execute("INSERT INTO executors (name) VALUES (?)", (name,))
                    ids.append(cur.lastrowid)
            conn.execute("UPDATE items SET executors_json = ? WHERE id = ?",
                         (json.dumps(ids), row["id"]))
        except (ValueError, TypeError, IndexError):
            pass

    conn.commit()
    conn.close()


init_db()


def row_to_dict(row):
    return dict(row)


def resolve_executors(conn, executors_json: str) -> list:
    """
    Превращает JSON-строку с массивом ID исполнителей в список объектов
    {id, name, department_name}. Поддерживает устаревший формат (строки-имена).
    """
    try:
        data = json.loads(executors_json) if executors_json else []
    except (ValueError, TypeError):
        return []

    result = []
    for item in data:
        if isinstance(item, int):
            row = conn.execute("""
                SELECT e.id, e.name, d.name AS department_name
                FROM executors e
                LEFT JOIN departments d ON e.department_id = d.id
                WHERE e.id = ?
            """, (item,)).fetchone()
            if row:
                result.append(row_to_dict(row))
            else:
                # Исполнитель удалён из справочника, но привязка к задаче
                # сохраняется (ID не теряется при последующем сохранении).
                result.append({"id": item, "name": f"#{item} (удалён)", "department_name": None})
        elif isinstance(item, str) and item.strip():
            result.append({"id": None, "name": item, "department_name": None})
    return result


# ── Отделы ────────────────────────────────────────────────────────────────────

class DepartmentCreate(BaseModel):
    name: str


@app.get("/api/departments")
def list_departments():
    conn = get_db()
    rows = conn.execute("SELECT * FROM departments ORDER BY name").fetchall()
    conn.close()
    return [row_to_dict(r) for r in rows]


@app.post("/api/departments", status_code=201)
def create_department(body: DepartmentCreate):
    name = body.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Name required")
    conn = get_db()
    if conn.execute("SELECT id FROM departments WHERE name = ?", (name,)).fetchone():
        conn.close()
        raise HTTPException(status_code=409, detail="Already exists")
    cur = conn.execute("INSERT INTO departments (name) VALUES (?)", (name,))
    conn.commit()
    row = conn.execute("SELECT * FROM departments WHERE id = ?", (cur.lastrowid,)).fetchone()
    conn.close()
    return row_to_dict(row)


@app.put("/api/departments/{dept_id}")
def update_department(dept_id: int, body: DepartmentCreate):
    """Переименовывает отдел. 404 — если не найден, 409 — если имя занято другим."""
    name = body.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Name required")
    conn = get_db()
    if not conn.execute("SELECT id FROM departments WHERE id = ?", (dept_id,)).fetchone():
        conn.close()
        raise HTTPException(status_code=404, detail="Not found")
    if conn.execute("SELECT id FROM departments WHERE name = ? AND id != ?", (name, dept_id)).fetchone():
        conn.close()
        raise HTTPException(status_code=409, detail="Already exists")
    conn.execute("UPDATE departments SET name = ? WHERE id = ?", (name, dept_id))
    conn.commit()
    row = conn.execute("SELECT * FROM departments WHERE id = ?", (dept_id,)).fetchone()
    conn.close()
    return row_to_dict(row)


@app.delete("/api/departments/{dept_id}", status_code=204)
def delete_department(dept_id: int):
    """Удаляет отдел; у исполнителей этого отдела department_id становится NULL."""
    conn = get_db()
    if not conn.execute("SELECT id FROM departments WHERE id = ?", (dept_id,)).fetchone():
        conn.close()
        raise HTTPException(status_code=404, detail="Not found")
    conn.execute("UPDATE executors SET department_id = NULL WHERE department_id = ?", (dept_id,))
    conn.execute("DELETE FROM departments WHERE id = ?", (dept_id,))
    conn.commit()
    conn.close()


# ── Исполнители ───────────────────────────────────────────────────────────────

class ExecutorCreate(BaseModel):
    name: str
    department_id: Optional[int] = None


@app.get("/api/executors")
def list_executors():
    """Возвращает всех исполнителей с названием отдела, сортировка: отдел → имя."""
    conn = get_db()
    rows = conn.execute("""
        SELECT e.id, e.name, e.department_id, d.name AS department_name
        FROM executors e
        LEFT JOIN departments d ON e.department_id = d.id
        ORDER BY d.name, e.name
    """).fetchall()
    conn.close()
    return [row_to_dict(r) for r in rows]


@app.post("/api/executors", status_code=201)
def create_executor(body: ExecutorCreate):
    name = body.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Name required")
    conn = get_db()
    if conn.execute("SELECT id FROM executors WHERE name = ?", (name,)).fetchone():
        conn.close()
        raise HTTPException(status_code=409, detail="Already exists")
    cur = conn.execute("INSERT INTO executors (name, department_id) VALUES (?, ?)",
                       (name, body.department_id))
    conn.commit()
    row = conn.execute("""
        SELECT e.id, e.name, e.department_id, d.name AS department_name
        FROM executors e LEFT JOIN departments d ON e.department_id = d.id
        WHERE e.id = ?
    """, (cur.lastrowid,)).fetchone()
    conn.close()
    return row_to_dict(row)


@app.put("/api/executors/{executor_id}")
def update_executor(executor_id: int, body: ExecutorCreate):
    """Меняет имя и/или отдел исполнителя. 404 — если нет, 409 — если имя занято другим."""
    name = body.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Name required")
    conn = get_db()
    if not conn.execute("SELECT id FROM executors WHERE id = ?", (executor_id,)).fetchone():
        conn.close()
        raise HTTPException(status_code=404, detail="Not found")
    if conn.execute("SELECT id FROM executors WHERE name = ? AND id != ?", (name, executor_id)).fetchone():
        conn.close()
        raise HTTPException(status_code=409, detail="Already exists")
    conn.execute("UPDATE executors SET name = ?, department_id = ? WHERE id = ?",
                 (name, body.department_id, executor_id))
    conn.commit()
    row = conn.execute("""
        SELECT e.id, e.name, e.department_id, d.name AS department_name
        FROM executors e LEFT JOIN departments d ON e.department_id = d.id
        WHERE e.id = ?
    """, (executor_id,)).fetchone()
    conn.close()
    return row_to_dict(row)


@app.delete("/api/executors/{executor_id}", status_code=204)
def delete_executor(executor_id: int):
    """Удаляет исполнителя из справочника. Задачи не затрагиваются."""
    conn = get_db()
    if not conn.execute("SELECT id FROM executors WHERE id = ?", (executor_id,)).fetchone():
        conn.close()
        raise HTTPException(status_code=404, detail="Not found")
    conn.execute("DELETE FROM executors WHERE id = ?", (executor_id,))
    conn.commit()
    conn.close()


# ── Задачи ────────────────────────────────────────────────────────────────────

class ItemCreate(BaseModel):
    topic: str
    ticket: Optional[str] = None
    priority: Optional[str] = None
    state: str = "in_progress"
    due_date: Optional[str] = None
    executors: list[int] = []        # список ID исполнителей
    status_date: Optional[str] = None
    status_note: Optional[str] = None


class ItemUpdate(BaseModel):
    topic: Optional[str] = None
    ticket: Optional[str] = None
    priority: Optional[str] = None
    state: Optional[str] = None
    due_date: Optional[str] = None
    executors: Optional[list[int]] = None


def serialize_item(conn, item):
    """Собирает полный словарь задачи: исполнители + 3 последних статуса + счётчик."""
    d = row_to_dict(item)
    ej = d.pop("executors_json", None)
    d["executors"] = resolve_executors(conn, ej)
    recent = conn.execute(
        "SELECT * FROM statuses WHERE item_id = ? ORDER BY status_date DESC, id DESC LIMIT 3",
        (item["id"],),
    ).fetchall()
    recent_list = [row_to_dict(r) for r in recent]
    d["recent_statuses"] = recent_list
    # Общее число статусов — для счётчика «+N». Доп. запрос нужен,
    # только когда статусов может быть больше отданных трёх.
    if len(recent_list) < 3:
        d["status_count"] = len(recent_list)
    else:
        d["status_count"] = conn.execute(
            "SELECT COUNT(*) AS c FROM statuses WHERE item_id = ?", (item["id"],)
        ).fetchone()["c"]
    return d


@app.get("/api/items")
def list_items(state: Optional[str] = None):
    """Список задач с тремя последними статусами и полными данными исполнителей."""
    conn = get_db()
    if state:
        items = conn.execute("SELECT * FROM items WHERE state = ? ORDER BY id", (state,)).fetchall()
    else:
        items = conn.execute("SELECT * FROM items ORDER BY id").fetchall()
    result = [serialize_item(conn, item) for item in items]
    conn.close()
    return result


@app.get("/api/items/{item_id}")
def get_item(item_id: int):
    """Одна задача — для просмотра в отдельном окне (?task=ID)."""
    conn = get_db()
    item = conn.execute("SELECT * FROM items WHERE id = ?", (item_id,)).fetchone()
    if not item:
        conn.close()
        raise HTTPException(status_code=404, detail="Задача не найдена")
    d = serialize_item(conn, item)
    conn.close()
    return d


@app.post("/api/items", status_code=201)
def create_item(item: ItemCreate):
    conn = get_db()
    ej = json.dumps(item.executors) if item.executors else None
    cur = conn.execute(
        "INSERT INTO items (topic, ticket, priority, state, due_date, executors_json) VALUES (?, ?, ?, ?, ?, ?)",
        (item.topic, item.ticket, item.priority, item.state, item.due_date, ej),
    )
    item_id = cur.lastrowid
    if item.status_note:
        conn.execute(
            "INSERT INTO statuses (item_id, status_date, status_note) VALUES (?, ?, ?)",
            (item_id, item.status_date, item.status_note),
        )
    conn.commit()
    row = conn.execute("SELECT * FROM items WHERE id = ?", (item_id,)).fetchone()
    d = row_to_dict(row)
    ej = d.pop("executors_json", None)
    d["executors"] = resolve_executors(conn, ej)
    conn.close()
    return d


@app.put("/api/items/{item_id}")
def update_item(item_id: int, item: ItemUpdate):
    conn = get_db()
    if not conn.execute("SELECT id FROM items WHERE id = ?", (item_id,)).fetchone():
        conn.close()
        raise HTTPException(status_code=404, detail="Item not found")

    required   = {"topic", "state"}
    raw_fields = item.model_dump(exclude_unset=True)
    fields     = {}

    for k, v in raw_fields.items():
        if k == "executors":
            # Пустой список → NULL в БД (очищает исполнителей)
            fields["executors_json"] = json.dumps(v) if v else None
        elif k in required and not v:
            pass  # не обнуляем обязательные поля
        else:
            fields[k] = v

    if fields:
        set_clause = ", ".join(f"{k} = ?" for k in fields)
        conn.execute(f"UPDATE items SET {set_clause} WHERE id = ?", list(fields.values()) + [item_id])
        conn.commit()

    row = conn.execute("SELECT * FROM items WHERE id = ?", (item_id,)).fetchone()
    d = row_to_dict(row)
    ej = d.pop("executors_json", None)
    d["executors"] = resolve_executors(conn, ej)
    conn.close()
    return d


@app.delete("/api/items/{item_id}", status_code=204)
def delete_item(item_id: int):
    conn = get_db()
    if not conn.execute("SELECT id FROM items WHERE id = ?", (item_id,)).fetchone():
        conn.close()
        raise HTTPException(status_code=404, detail="Item not found")
    conn.execute("DELETE FROM statuses WHERE item_id = ?", (item_id,))
    conn.execute("DELETE FROM items WHERE id = ?", (item_id,))
    conn.commit()
    conn.close()


# ── Статусы ───────────────────────────────────────────────────────────────────

class StatusCreate(BaseModel):
    status_date: Optional[str] = None
    status_note: Optional[str] = None


@app.get("/api/items/{item_id}/statuses")
def list_statuses(item_id: int):
    conn = get_db()
    if not conn.execute("SELECT id FROM items WHERE id = ?", (item_id,)).fetchone():
        conn.close()
        raise HTTPException(status_code=404, detail="Item not found")
    rows = conn.execute(
        "SELECT * FROM statuses WHERE item_id = ? ORDER BY status_date DESC, id DESC",
        (item_id,),
    ).fetchall()
    conn.close()
    return [row_to_dict(r) for r in rows]


@app.post("/api/items/{item_id}/statuses", status_code=201)
def add_status(item_id: int, body: StatusCreate):
    conn = get_db()
    if not conn.execute("SELECT id FROM items WHERE id = ?", (item_id,)).fetchone():
        conn.close()
        raise HTTPException(status_code=404, detail="Item not found")
    cur = conn.execute(
        "INSERT INTO statuses (item_id, status_date, status_note) VALUES (?, ?, ?)",
        (item_id, body.status_date, body.status_note),
    )
    conn.commit()
    row = conn.execute("SELECT * FROM statuses WHERE id = ?", (cur.lastrowid,)).fetchone()
    conn.close()
    return row_to_dict(row)


@app.delete("/api/statuses/{status_id}", status_code=204)
def delete_status(status_id: int):
    conn = get_db()
    if not conn.execute("SELECT id FROM statuses WHERE id = ?", (status_id,)).fetchone():
        conn.close()
        raise HTTPException(status_code=404, detail="Status not found")
    conn.execute("DELETE FROM statuses WHERE id = ?", (status_id,))
    conn.commit()
    conn.close()


# ── Экспорт / Импорт ──────────────────────────────────────────────────────────

CSV_FIELDS = ["id", "topic", "ticket", "priority", "state", "due_date",
              "executors", "last_status_date", "last_status_note"]


@app.get("/api/export/csv")
def export_csv():
    """Экспортирует задачи в CSV. Исполнители — 'Отдел — Имя', разделитель |."""
    conn = get_db()
    items = conn.execute("SELECT * FROM items ORDER BY id").fetchall()

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=CSV_FIELDS)
    writer.writeheader()

    for item in items:
        last  = conn.execute(
            "SELECT * FROM statuses WHERE item_id = ? ORDER BY status_date DESC, id DESC LIMIT 1",
            (item["id"],),
        ).fetchone()
        execs = resolve_executors(conn, item["executors_json"])
        exec_str = " | ".join(
            f"{e['department_name']} — {e['name']}" if e.get("department_name") else e["name"]
            for e in execs
        )
        writer.writerow({
            "id":               item["id"],
            "topic":            item["topic"],
            "ticket":           item["ticket"] or "",
            "priority":         item["priority"] or "",
            "state":            item["state"],
            "due_date":         item["due_date"] or "",
            "executors":        exec_str,
            "last_status_date": last["status_date"] or "" if last else "",
            "last_status_note": last["status_note"] or "" if last else "",
        })

    conn.close()
    content = ("﻿" + output.getvalue()).encode("utf-8")
    return StreamingResponse(
        io.BytesIO(content),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=protocol.csv"},
    )


@app.get("/api/export/xlsx")
def export_xlsx():
    """Экспортирует задачи в Excel (.xlsx) с форматированием."""
    conn = get_db()
    items = conn.execute("SELECT * FROM items ORDER BY id").fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Протокол"

    # ── Стили ──────────────────────────────────────────────────────────────────
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_font    = Font(bold=True, color="FFFFFF", size=11)
    header_fill    = PatternFill("solid", fgColor="1E3A5F")
    header_align   = Alignment(horizontal="center", vertical="center", wrap_text=True)

    PRIO_FILL = {
        "high":   PatternFill("solid", fgColor="FEE2E2"),
        "medium": PatternFill("solid", fgColor="FEF9C3"),
        "low":    PatternFill("solid", fgColor="DCFCE7"),
    }
    STATE_FILL = {
        "in_progress": PatternFill("solid", fgColor="DBEAFE"),
        "postponed":   PatternFill("solid", fgColor="F3F4F6"),
        "closed":      PatternFill("solid", fgColor="D1FAE5"),
    }
    STATE_LABEL_RU = {
        "in_progress": "В работе",
        "postponed":   "Отложено",
        "closed":      "Закрыто",
    }
    PRIO_LABEL_RU = {
        "high":   "Высокий",
        "medium": "Средний",
        "low":    "Низкий",
        "":       "",
    }

    # ── Заголовки ─────────────────────────────────────────────────────────────
    headers = ["№", "Тема", "Тикет", "Приоритет", "Состояние", "Срок",
               "Исполнители", "Дата статуса", "Последний статус"]
    col_widths = [5, 45, 12, 12, 12, 12, 30, 14, 50]

    for col_idx, (hdr, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=hdr)
        cell.font   = header_font
        cell.fill   = header_fill
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # ── Данные: предзагрузка за 2 запроса вместо N+1 ─────────────────────────
    # Последний статус для каждой задачи (ROW_NUMBER — SQLite ≥ 3.25)
    last_statuses: dict = {}
    for r in conn.execute("""
        SELECT * FROM (
            SELECT *, ROW_NUMBER() OVER (
                PARTITION BY item_id ORDER BY status_date DESC, id DESC
            ) AS rn FROM statuses
        ) WHERE rn = 1
    """).fetchall():
        last_statuses[r["item_id"]] = row_to_dict(r)

    # Все исполнители с отделами — один запрос
    all_executors: dict = {}
    for r in conn.execute("""
        SELECT e.id, e.name, d.name AS department_name
        FROM executors e LEFT JOIN departments d ON d.id = e.department_id
    """).fetchall():
        all_executors[r["id"]] = row_to_dict(r)

    center = Alignment(horizontal="center", vertical="top")
    wrap   = Alignment(vertical="top", wrap_text=True)
    top    = Alignment(vertical="top")

    for row_idx, item in enumerate(items, 2):
        last = last_statuses.get(item["id"])
        exec_ids = json.loads(item["executors_json"]) if item["executors_json"] else []
        execs = [all_executors[eid] for eid in exec_ids if eid in all_executors]
        exec_str = " | ".join(
            f"{e['department_name']} — {e['name']}" if e.get("department_name") else e["name"]
            for e in execs
        )

        prio  = item["priority"] or ""
        state = item["state"]
        row_data = [
            row_idx - 1,
            item["topic"],
            item["ticket"] or "",
            PRIO_LABEL_RU.get(prio, prio),
            STATE_LABEL_RU.get(state, state),
            item["due_date"] or "",
            exec_str,
            last["status_date"] or "" if last else "",
            last["status_note"] or "" if last else "",
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = wrap if col_idx in (2, 7, 9) else (center if col_idx in (1, 3, 4, 5, 6, 8) else top)

        # цвет по приоритету (столбец 4) и состоянию (столбец 5)
        if prio in PRIO_FILL:
            ws.cell(row=row_idx, column=4).fill = PRIO_FILL[prio]
        if state in STATE_FILL:
            ws.cell(row=row_idx, column=5).fill = STATE_FILL[state]

        ws.row_dimensions[row_idx].height = 20

    conn.close()

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=protocol.xlsx"},
    )


@app.post("/api/import/csv")
async def import_csv(file: UploadFile = File(...)):
    """
    Импортирует задачи из CSV. Исполнители сопоставляются по имени;
    если не найден — создаётся новый без отдела.
    """
    raw  = await file.read()
    text = raw.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))

    conn = get_db()
    imported = 0
    for row in reader:
        topic = (row.get("topic") or "").strip()
        if not topic:
            continue

        # Разбираем исполнителей из строки вида "Отдел — Имя | Имя2"
        exec_ids = []
        for part in (row.get("executors") or "").split("|"):
            part = part.strip()
            if not part:
                continue
            name = part.split(" — ")[-1].strip()  # берём часть после " — " или всю строку
            existing = conn.execute("SELECT id FROM executors WHERE name = ?", (name,)).fetchone()
            if existing:
                exec_ids.append(existing["id"])
            else:
                cur2 = conn.execute("INSERT INTO executors (name) VALUES (?)", (name,))
                exec_ids.append(cur2.lastrowid)

        ej = json.dumps(exec_ids) if exec_ids else None
        cur = conn.execute(
            "INSERT INTO items (topic, ticket, priority, state, due_date, executors_json) VALUES (?, ?, ?, ?, ?, ?)",
            (topic, row.get("ticket") or None, row.get("priority") or None,
             row.get("state") or "in_progress", row.get("due_date") or None, ej),
        )
        s_note = row.get("last_status_note") or None
        if s_note:
            conn.execute(
                "INSERT INTO statuses (item_id, status_date, status_note) VALUES (?, ?, ?)",
                (cur.lastrowid, row.get("last_status_date") or None, s_note),
            )
        imported += 1

    conn.commit()
    conn.close()
    return {"imported": imported}


app.mount("/", StaticFiles(directory="static", html=True), name="static")
